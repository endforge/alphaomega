"""
File: text_extractor.py

Purpose:
    Produces canonical text content from supported source objects.

This module contains format-specific text extraction behavior.
It does not retrieve source objects, determine synchronization state,
persist Knowledge Objects, or perform AI processing.
"""

import io
from pathlib import Path

from bs4 import BeautifulSoup
from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader


class TextExtractor:
    """
    Extract canonical text content from supported file formats.
    """

    extractor_name = "text_extractor"

    supported_extensions = {
        ".txt",
        ".md",
        ".csv",
        ".json",
        ".xml",
        ".docx",
        ".pdf",
        ".xlsx",
        ".html",
        ".htm",
    }

    @classmethod
    def supports(cls, file_name):
        """
        Return True when the file extension is supported.
        """

        extension = Path(file_name).suffix.lower()

        return extension in cls.supported_extensions

    @classmethod
    def extract(cls, file_name, file_bytes):
        """
        Extract canonical text from a supported source object.

        Args:
            file_name:
                Source object filename.

            file_bytes:
                Raw source object content.

        Returns:
            str:
                Extracted canonical text.

        Raises:
            ValueError:
                If the source object's file type is unsupported.
        """

        extension = Path(file_name).suffix.lower()

        if extension in {
            ".txt",
            ".md",
            ".csv",
            ".json",
            ".xml",
        }:
            return cls._extract_txt(file_bytes)

        if extension == ".docx":
            return cls._extract_docx(file_bytes)

        if extension == ".pdf":
            return cls._extract_pdf(file_bytes)

        if extension == ".xlsx":
            return cls._extract_xlsx(file_bytes)

        if extension in {".html", ".htm"}:
            return cls._extract_html(file_bytes)

        raise ValueError(
            f"Unsupported extraction format: '{extension}'."
        )

    @staticmethod
    def _extract_txt(file_bytes):
        """
        Extract UTF-8 text.
        """

        return file_bytes.decode(
            "utf-8",
            errors="ignore",
        )

    @staticmethod
    def _extract_docx(file_bytes):
        """
        Extract non-empty paragraphs from a DOCX document.
        """

        document = Document(
            io.BytesIO(file_bytes)
        )

        return "\n".join(
            paragraph.text
            for paragraph in document.paragraphs
            if paragraph.text.strip()
        )

    @staticmethod
    def _extract_pdf(file_bytes):
        """
        Extract text from populated PDF pages.
        """

        reader = PdfReader(
            io.BytesIO(file_bytes)
        )

        pages = []

        for page in reader.pages:
            text = page.extract_text() or ""

            if text.strip():
                pages.append(text)

        return "\n".join(pages)

    @staticmethod
    def _extract_xlsx(file_bytes):
        """
        Extract populated worksheet cells as canonical text.
        """

        workbook = load_workbook(
            io.BytesIO(file_bytes),
            read_only=True,
            data_only=True,
        )

        rows = []

        for sheet in workbook.worksheets:
            rows.append(
                f"\n--- Sheet: {sheet.title} ---"
            )

            for row in sheet.iter_rows(
                values_only=True
            ):
                values = [
                    str(cell)
                    for cell in row
                    if cell is not None
                ]

                if values:
                    rows.append(
                        " | ".join(values)
                    )

        return "\n".join(rows)

    @staticmethod
    def _extract_html(file_bytes):
        """
        Extract visible text from HTML content.
        """

        soup = BeautifulSoup(
            file_bytes,
            "html.parser",
        )

        return soup.get_text(
            separator="\n"
        )