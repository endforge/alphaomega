import io
import json
import requests
from msal import PublicClientApplication
from pathlib import Path
from docx import Document
from pypdf import PdfReader
from openpyxl import load_workbook
from bs4 import BeautifulSoup

CLIENT_ID = "1ec08875-5201-4c26-8b61-97c0916e0885"

# -----------------------------
# File paths
# -----------------------------

MANIFEST_PATH = Path("data/manifest/onedrive_manifest.json")
OUTPUT_DIR = Path("data/extracted_text")

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# -----------------------------
# Get Access Token
# -----------------------------

def get_access_token():
    app = PublicClientApplication(
        CLIENT_ID,
        authority="https://login.microsoftonline.com/common"
    )

    result = app.acquire_token_interactive(
        scopes=["User.Read", "Files.Read"]
    )

    return result["access_token"]

# -----------------------------
# OneDrive download
# -----------------------------

def download_file_bytes(access_token, item_id):
    url = f"https://graph.microsoft.com/v1.0/me/drive/items/{item_id}/content"

    headers = {
        "Authorization": f"Bearer {access_token}"
    }

    response = requests.get(url, headers=headers)
    response.raise_for_status()

    return response.content


# -----------------------------
# Text extraction functions
# -----------------------------

def extract_txt(file_bytes):
    return file_bytes.decode("utf-8", errors="ignore")


def extract_docx(file_bytes):
    doc = Document(io.BytesIO(file_bytes))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def extract_pdf(file_bytes):
    reader = PdfReader(io.BytesIO(file_bytes))

    pages = []

    for page in reader.pages:
        text = page.extract_text() or ""
        if text.strip():
            pages.append(text)

    return "\n".join(pages)


def extract_xlsx(file_bytes):
    workbook = load_workbook(
        io.BytesIO(file_bytes),
        read_only=True,
        data_only=True
    )

    rows = []

    for sheet in workbook.worksheets:
        rows.append(f"\n--- Sheet: {sheet.title} ---")

        for row in sheet.iter_rows(values_only=True):
            values = [str(cell) for cell in row if cell is not None]

            if values:
                rows.append(" | ".join(values))

    return "\n".join(rows)


def extract_html(file_bytes):
    soup = BeautifulSoup(file_bytes, "html.parser")
    return soup.get_text(separator="\n")


# -----------------------------
# File type router
# -----------------------------

def extract_text(file_name, file_bytes):
    extension = Path(file_name).suffix.lower()

    if extension in [".txt", ".md", ".csv", ".json", ".xml"]:
        return extract_txt(file_bytes)

    if extension == ".docx":
        return extract_docx(file_bytes)

    if extension == ".pdf":
        return extract_pdf(file_bytes)

    if extension == ".xlsx":
        return extract_xlsx(file_bytes)

    if extension in [".html", ".htm"]:
        return extract_html(file_bytes)

    return None


# -----------------------------
# Save extracted output
# -----------------------------

def save_extracted_text(item, text):
    safe_id = item["id"].replace("!", "_")

    text_path = OUTPUT_DIR / f"{safe_id}.txt"
    metadata_path = OUTPUT_DIR / f"{safe_id}.json"

    text_path.write_text(text, encoding="utf-8")

    metadata = {
        "id": item["id"],
        "name": item.get("name"),
        "path": item.get("path"),
        "extension": Path(item.get("name", "")).suffix.lower(),
        "size": item.get("size"),
        "modified": item.get("last_modified")
    }

    metadata_path.write_text(
        json.dumps(metadata, indent=2),
        encoding="utf-8"
    )


# -----------------------------
# Load manifest
# -----------------------------

def load_manifest():
    return json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))


# -----------------------------
# Main processing loop
# -----------------------------

def process_manifest(access_token):
    manifest = load_manifest()

    processed = 0
    skipped = 0
    failed = 0

    for item in manifest:
        file_name = item.get("name", "")
        item_id = item.get("id")

        if not item_id or not file_name:
            skipped += 1
            continue

        extension = Path(file_name).suffix.lower()

        if extension not in [
            ".txt", ".md", ".csv", ".json", ".xml",
            ".docx", ".pdf", ".xlsx",
            ".html", ".htm"
        ]:
            skipped += 1
            continue

        try:
            print(f"Processing: {file_name}")

            file_bytes = download_file_bytes(access_token, item_id)
            text = extract_text(file_name, file_bytes)

            if text and text.strip():
                save_extracted_text(item, text)
                processed += 1
            else:
                skipped += 1

        except Exception as error:
            failed += 1
            print(f"FAILED: {file_name}")
            print(f"Reason: {error}")

    print("\nExtraction complete.")
    print(f"Processed: {processed}")
    print(f"Skipped:   {skipped}")
    print(f"Failed:    {failed}")


# -----------------------------
# Script entry point
# -----------------------------

if __name__ == "__main__":
    access_token = get_access_token()
    process_manifest(access_token)